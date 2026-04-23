# -*- coding: utf-8 -*-
"""
Сводный прайс: загрузка нескольких Excel-прайсов, извлечение данных по производителю/препаратам,
формирование матрицы (препараты × дистрибьюторы).
"""
from __future__ import annotations

import argparse
import ctypes
import re
import shutil
import sys
import threading
import time
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rapidfuzz import fuzz

# --- константы ---
DEFAULT_KEY_CANDIDATES = ["Артикул", "Код", "SKU", "ID", "PartNumber", "НоменклатураКод"]
DEFAULT_NAME_CANDIDATES = [
    "Наименование",
    "Наименование товаров",
    "Наименование товара",
    "Наименование препарата",
    "Наименование препаратов",
    "Товары",
    "Махсулот номи",
    "Номенклатура",
    "Товар",
    "Название",
    "Name",
    "Наимменование",
]
DEFAULT_PRICE_CANDIDATES = [
    "Цена (100%) с НДС",
    "Цена При 100 % оплате цена",
    "При 100 % оплате цена",
    "Тип цены Сотув",
    "Цена С НДС",
    "Цена 100%",
    "Цена с НДС 100%",
    "Цена для клиента с предоплатой 100%",
    "Цена с НДС",
    "Цена без НДС",
    "Цена",
    "Цена Спец мин 1млн",
    "Тип цены При 100% Оплате (перечисление)",
    "При 100% Оплате",
    "Цена договорная",
    "Цена (с НДС)",
    "Цена Договорная",
    "Цена переч",
    "Цена переч:",
    "Цена реал",
    "ПриходЦена",
    "Сумма заказ",
    "Сумма",
    "Отп. Цена",
    "Цена, руб",
    "Price",
    "Нарх",
    "Розница",
    "Опт",
]
DEFAULT_STOCK_CANDIDATES = ["Остаток", "Наличие", "Stock", "Qty", "Количество"]
DEFAULT_PRODUCER_CANDIDATES = ["Производитель", "Бренд", "Manufacturer", "Ишлаб чикарувчи"]
DEFAULT_EXPIRY_CANDIDATES = ["Срок годности", "Срок", "Годен до", "Expiry"]


def _normalize_col_for_match(col: str) -> str:
    """Нормализация названия колонки для сопоставления с кандидатами."""
    if not isinstance(col, str):
        return ""
    s = col.replace("\n", " ").replace("\r", " ").replace("\t", " ").strip()
    while s and s[-1] in ":;":
        s = s[:-1].strip()
    # Регистронезависимо: в части .xls заголовки приходят строчными (наимменование, цена).
    return s.casefold()


def _clean_col(c: Any) -> str:
    return str(c).strip() if c is not None else ""


def pick_column(
    columns: List[str],
    candidates: List[str],
) -> Optional[str]:
    """Выбирает первую колонку из df, нормализованное имя которой совпадает с одним из кандидатов."""
    norm_cols = {_normalize_col_for_match(c): c for c in columns}
    for cand in candidates:
        n = _normalize_col_for_match(cand)
        if n in norm_cols:
            return norm_cols[n]
    return None


def pick_price_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    """
    Выбирает колонку цены:
    1) сначала точное совпадение по candidates;
    2) если найдено несколько, приоритет у колонок с "цена/price/нарх",
       а "сумма" считается менее приоритетной.
    """
    norm_cols = {_normalize_col_for_match(c): c for c in columns}
    present = []
    for idx, cand in enumerate(candidates):
        n = _normalize_col_for_match(cand)
        if n in norm_cols:
            present.append((idx, n, norm_cols[n]))
    if not present:
        return None

    def score(name_norm: str, pos: int) -> Tuple[int, int]:
        low = name_norm.lower()
        strong = ("цена" in low) or ("price" in low) or ("нарх" in low)
        weak_sum = "сумма" in low
        # Более высокий score -> выше приоритет. pos инвертирован, чтобы раньше в candidates было лучше.
        return (2 if strong else 0) + (0 if weak_sum else 1), -pos

    best = max(present, key=lambda x: score(x[1], x[0]))
    return best[2]


def _excel_engine_for(path: Path) -> str:
    return "openpyxl" if path.suffix.lower() == ".xlsx" else "xlrd"


def _grand_farm_header_row(
    path: Path,
    name_candidates: List[str],
    price_candidates: List[str],
) -> Optional[int]:
    """
    В прайсах «Гранд Фарм» таблица начинается на сотнях строки; номер строки
    заголовков в файле со временем меняется. Ищем первую подходящую строку
    в диапазоне Excel (read_only, без полного парса через pandas).
    Возвращает индекс для pandas header= (0-based), как в read_excel(header=...).
    """
    price_cands = {_normalize_col_for_match(p) for p in price_candidates}
    name_cands = {_normalize_col_for_match(n) for n in name_candidates}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for excel_row_1based, row in enumerate(
            ws.iter_rows(min_row=250, max_row=620, values_only=True),
            start=250,
        ):
            normed: set = set()
            for v in row:
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                normed.add(_normalize_col_for_match(s))
            if normed & name_cands and normed & price_cands:
                return excel_row_1based - 1
    finally:
        wb.close()
    return None


def read_price_table(
    path: Path,
    key_candidates: List[str],
    name_candidates: List[str],
    price_candidates: List[str],
    stock_candidates: List[str],
    use_key: bool,
    use_name: bool,
    read_preset: Optional[str] = None,
) -> pd.DataFrame:
    """
    Читает прайс даже если заголовок не в первой строке.
    Спец-обработка: Хожи Акбар Фарм — заголовки на 12-й строке (header=11).
    read_preset: None (авто по имени файла), \"khoji_akbar\", \"grand_farm\".
    """
    engine = _excel_engine_for(path)
    read_path = path

    use_khoji = read_preset == "khoji_akbar" or (
        read_preset is None and "Хожи Акбар Фарм" in path.stem
    )
    use_grand = read_preset == "grand_farm" or (
        read_preset is None and "Гранд Фарм" in path.stem and engine == "openpyxl"
    )

    if use_khoji:
        try:
            df = pd.read_excel(path, engine=engine, header=11)
        except PermissionError:
            out_dir = path.parent / "output" / "_tmp_read"
            out_dir.mkdir(parents=True, exist_ok=True)
            read_path = out_dir / path.name
            shutil.copy2(path, read_path)
            df = pd.read_excel(read_path, engine=engine, header=11)
        df.columns = [_clean_col(c) for c in df.columns]
        return df

    if use_grand:
        def _read_grand(headers: int) -> pd.DataFrame:
            return pd.read_excel(read_path, engine=engine, header=headers)

        hdr: Optional[int] = None
        try:
            hdr = _grand_farm_header_row(path, name_candidates, price_candidates)
        except Exception:
            hdr = None
        if hdr is None:
            hdr = 402
        try:
            df = _read_grand(hdr)
        except PermissionError:
            out_dir = path.parent / "output" / "_tmp_read"
            out_dir.mkdir(parents=True, exist_ok=True)
            read_path = out_dir / path.name
            shutil.copy2(path, read_path)
            df = _read_grand(hdr)
        df.columns = [_clean_col(c) for c in df.columns]
        if pick_column(list(df.columns), name_candidates) and pick_price_column(
            list(df.columns), price_candidates
        ):
            return df
        try:
            hdr2 = _grand_farm_header_row(read_path, name_candidates, price_candidates)
        except Exception:
            hdr2 = None
        if hdr2 is not None and hdr2 != hdr:
            df = pd.read_excel(read_path, engine=engine, header=hdr2)
            df.columns = [_clean_col(c) for c in df.columns]
            if pick_column(list(df.columns), name_candidates) and pick_price_column(
                list(df.columns), price_candidates
            ):
                return df
        # не удалось — падаем в общий поиск заголовка ниже

    max_scan = 600
    price_cands = [_normalize_col_for_match(p) for p in price_candidates]
    name_cands = [_normalize_col_for_match(n) for n in name_candidates]

    # Быстрый путь: пробуем обычный header=0.
    try:
        df0 = pd.read_excel(read_path, engine=engine, header=0)
        df0.columns = [_clean_col(c) for c in df0.columns]
        cols0 = [_normalize_col_for_match(c) for c in df0.columns]
        if any(p in cols0 for p in price_cands) and any(n in cols0 for n in name_cands):
            return df0
    except PermissionError:
        out_dir = path.parent / "output" / "_tmp_read"
        out_dir.mkdir(parents=True, exist_ok=True)
        read_path = out_dir / path.name
        shutil.copy2(path, read_path)
    except Exception:
        pass

    # Медленный путь: ищем строку заголовков, но читаем только 0 строк (только названия колонок).
    for header_row in range(max_scan):
        try:
            probe = pd.read_excel(read_path, engine=engine, header=header_row, nrows=0)
        except PermissionError:
            out_dir = path.parent / "output" / "_tmp_read"
            out_dir.mkdir(parents=True, exist_ok=True)
            read_path = out_dir / path.name
            shutil.copy2(path, read_path)
            probe = pd.read_excel(read_path, engine=engine, header=header_row, nrows=0)
        except Exception:
            continue

        if len(probe.columns) < 2:
            continue

        row_vals = [_normalize_col_for_match(str(c)) for c in probe.columns]
        has_price = any(p in row_vals for p in price_cands)
        has_name = any(n in row_vals for n in name_cands)
        if not (has_price and has_name):
            continue

        df = pd.read_excel(read_path, engine=engine, header=header_row)
        df.columns = [_clean_col(c) for c in df.columns]
        return df

    raise ValueError(f"Не удалось найти строку заголовков с ценой и наименованием в файле: {path}")


def to_number(val: Any) -> Optional[float]:
    """Преобразует значение ячейки в число (цену)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".").replace(" ", "")
    for ch in ["\xa0", "₽", "руб", "руб.", "сум"]:
        s = s.replace(ch, "")
    s = s.strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def round_price(val: float) -> Optional[int]:
    """Округляет цену до целого по правилу .5 вверх."""
    try:
        return int(Decimal(str(val)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError, TypeError):
        return None


def normalize_name(s: str) -> str:
    """Нормализация названия для сравнения."""
    if not s or not isinstance(s, str):
        return ""
    return " ".join(str(s).lower().split())


# Суффиксы источника/производителя, которые встречаются в конце названия
# в прайсах («климаксан 10г/мат.мед», «аназин-к/мат»).
_SOURCE_SUFFIX_RE = re.compile(
    r"[/\\]\s*(?:мат\.?\s*мед\.?|мат\.?медика?|materia?|mm|мм)\s*$",
    re.IGNORECASE,
)


# Таблица нормализации фармацевтических форм выпуска.
# ВАЖНО: паттерны от длинного к короткому — «гранулы» совпадёт раньше «гран».
# Все паттерны имеют \b в конце для точной границы слова.
_FORM_NORMALIZATION: List[Tuple[str, str]] = [
    # Гранулы (длинные формы сначала)
    (r"\bгранул[аыь]?\b", "гран"),
    (r"\bгран\b", "гран"),
    # Таблетки
    (r"\bтаблетк[аи]\b", "таб"),
    (r"\bтаблет\b", "таб"),
    (r"\bтабл\b", "таб"),
    (r"\bтаб\b", "таб"),
    # Капсулы
    (r"\bкапсул[аы]?\b", "кап"),
    (r"\bкапс\b", "кап"),
    (r"\bкап\b", "кап"),
    # Пастилки
    (r"\bпастилк[аи]\b", "паст"),
    (r"\bпаст\b", "паст"),
    # Граммы (единицы массы) — только как отдельное слово
    (r"\bграмм[а]?\b", "г"),
    # Миллиграммы
    (r"\bмиллиграм[м]?[а]?\b", "мг"),
    # Штуки / количество
    (r"\bштук[аи]?\b", "шт"),
    (r"\bшт\b", "шт"),
]


_STAR_MARKERS_RE = re.compile(r"[\*＊⁎∗⋆]+")


def strip_star_markers(s: str) -> str:
    """Убирает служебные звёздочки (например *** в прайсе Вероны) и похожие символы."""
    return _STAR_MARKERS_RE.sub("", s)


_TM_RE = re.compile(r"[®™©]+")


def strip_trademark_symbols(s: str) -> str:
    """Знаки ®, ™ и т.п. в прайсах не влияют на товар для сопоставления."""
    return _TM_RE.sub("", s)


# Склейка «Септа Назал» ↔ «СептаНазал»: два коротких кириллических слова подряд.
_NO_JOIN_TOKENS = frozenset(
    {
        "мг",
        "мл",
        "me",
        "таб",
        "кап",
        "спрей",
        "паст",
        "паста",
        "доз",
        "доза",
        "инъек",
        "для",
        "г",
        "дети",
        "взросл",
        "эвкалипт",
        "н",
    }
)


def _join_adjacent_cyrillic_name_parts(s: str) -> str:
    parts = s.split()
    if len(parts) < 2:
        return s
    out: List[str] = []
    i = 0
    while i < len(parts):
        a = parts[i]
        b = parts[i + 1] if i + 1 < len(parts) else None
        if (
            b is not None
            and re.fullmatch(r"[а-яё]{2,14}", a)
            and re.fullmatch(r"[а-яё]{2,14}", b)
            and len(a) <= 7
            and len(b) <= 7
            and len(a) + len(b) <= 11
            and a.casefold() not in _NO_JOIN_TOKENS
            and b.casefold() not in _NO_JOIN_TOKENS
        ):
            out.append(a + b)
            i += 2
        else:
            out.append(a)
            i += 1
    return " ".join(out)


def normalize_product_for_grouping(raw: str) -> str:
    """
    Нормализация названия для ключа группировки в сводном (режим «все по производителю»).
    Убирает типичные отличия между прайсами: звёздочки, скобки, пробелы в дозах, таб/табл., D/Д и т.д.
    Дополнительно нормализует словоформы форм выпуска (гранула/гранулы/гран → гран)
    и убирает суффиксы источника (/мат.мед и подобные).
    В отчёте по-прежнему показывается исходная строка (canonical_display).
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = str(raw).replace("\r", " ").replace("\n", " ")
    s = unicodedata.normalize("NFKC", s)
    # Убираем суффикс источника до всего остального (пока слэш ещё цел)
    s = _SOURCE_SUFFIX_RE.sub("", s).strip()
    s = normalize_name(s)
    s = strip_star_markers(s)
    s = strip_trademark_symbols(s)
    # «5мг+0,1мг» / плюс в комбинированных дозах
    s = s.replace("+", " ")
    # доза / доз
    s = re.sub(r"\bдоза\b", "доз", s)
    # спрей д/взрос / д/детей (Верона) vs полная концентрация в других прайсах
    s = re.sub(r"\bд/взрос\w*\b", "взросл ", s)
    s = re.sub(r"\bд/дет\w*\b", "дети ", s)
    # эвкалипт / Эвкалиптом в скобках или в тексте
    s = re.sub(r"\bэвкалипт\w*", "эвкалипт ", s)
    # «Септа Назал» → одно слово до разбора слэшей
    s = _join_adjacent_cyrillic_name_parts(s)
    s = re.sub(r"\s+", " ", s).strip()
    # слэш в дозах «80мг/12,5мг»
    s = s.replace("/", " ")
    # десятичные дозы: 12,5 / 12.5 → один токен 12d5 (точки потом вырежутся)
    s = re.sub(r"(\d)[.,](\d+)", r"\1d\2", s)
    # «5мг+0,1мг» / «0,1мг» в спреях СептаНазал — одна Нмн в разных прайсах
    if re.search(r"0d1", s):
        s = re.sub(r"(?:5мг|5\s+мг)\s+", "", s, count=1)
    if re.search(r"0d05", s):
        s = re.sub(r"(?:5мг|5\s+мг)\s+", "", s, count=1)
    # убрать пояснения в скобках — частая причина дублей (эвкалипт)/варианты
    s = re.sub(r"\[[^]]{0,120}]", " ", s)
    s = re.sub(r"\([^)]{0,120}\)", " ", s)
    # витамин D3 / Д3
    s = re.sub(r"\bвитамин\s+d\s*3\b", "витамин д3 ", s)
    s = re.sub(r"\bвитамин\s+д\s*3\b", "витамин д3 ", s)
    s = re.sub(r"витаминд3\b", "витамин д3", s)
    # 1000 ME / 1000МЕ (отдельным токеном me — иначе ломает правило «цифра+буква»)
    s = re.sub(r"(\d+)\s*(ме|me)\b", r"\1 me", s, flags=re.IGNORECASE)
    # H / Н в дозах: «таб. H 80мг», «Н80 80мг», «H160 160мг»
    s = re.sub(r"\bh\s*(\d+)", r" \1 ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bн\s*(\d+)", r" \1 ", s)
    s = re.sub(r"\bh(\d+)", r" \1 ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bн(\d+)", r" \1 ", s)
    # krka в наименовании
    s = re.sub(r"\bkrka\b", " ", s)
    # маркировка «-К» у торговых названий (гентамицин-к)
    s = re.sub(r"-к\b", " ", s)
    # п/о — часто только в одном прайсе
    s = re.sub(r"\bп\s*/\s*о\b", " ", s)
    # № / no / n + число
    s = s.replace("№", " n")
    s = re.sub(r"\bno\s*(\d)", r"n\1", s)
    # единая метка для инъекций (амп. / р-р д/ин. — одна позиция в прайсах)
    s = re.sub(r"\bамп\.?\b", "инъек ", s)
    s = re.sub(r"\bр-р\b", "инъек ", s)
    s = re.sub(r"\bд/ин\.?\b", "инъек ", s)
    s = re.sub(r"\bд/пр\.?\b", "инъек ", s)
    # --- Нормализация словоформ форм выпуска ---
    # Применяем до разбивки цифра/буква, чтобы «гран.10г» тоже попало.
    for pattern, replacement in _FORM_NORMALIZATION:
        s = re.sub(pattern, replacement + " ", s)
    # Убираем шумовые слова, которые есть только в части прайсов
    # (гомеоп, гомеопат — в одном прайсе, без них — в другом).
    s = re.sub(r"\bгомеоп\w*\b", " ", s, flags=re.IGNORECASE)

    s = re.sub(r"(\d)([а-яёa-z])", r"\1 \2", s, flags=re.IGNORECASE)
    s = re.sub(r"([а-яёa-z])(\d)", r"\1 \2", s, flags=re.IGNORECASE)
    # буквы (лат/кир), цифры, d в десятичных токенах (12d5), пробел
    s = re.sub(r"[^\d\sd\u0400-\u04FFa-z]", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    toks = [t for t in s.split() if t]
    if not toks:
        return ""
    # «н» отдельным токеном после нормализации H/Н
    toks = [t for t in toks if t.casefold() not in ("н", "h")]
    # «таб» при явной дозе / ME
    unit_like = any(
        t.casefold() in ("мг", "мл", "me")
        or t.casefold().endswith("me")
        or ("d" in t and any(c.isdigit() for c in t))
        for t in toks
    )
    if unit_like:
        toks = [t for t in toks if t.casefold() not in ("таб", "гран", "кап", "паст")]
    # Если есть токен «гран» — убираем голой «г» (дозировка в граммах),
    # чтобы «гранулы 10г» и «гранула 10 г» давали одинаковый ключ.
    if any(t.casefold() == "гран" for t in toks):
        toks = [t for t in toks if t.casefold() != "г"]
    # Дедупликация всех дублирующихся токенов (не только цифровых).
    # «гранула 10 гранулы» → гран, 10, гран → убираем второй гран.
    st = sorted(toks, key=str.casefold)
    deduped: List[str] = []
    for t in st:
        if deduped and deduped[-1].casefold() == t.casefold():
            continue
        deduped.append(t)
    return " ".join(deduped)


def fuzzy_merge_canonicals(
    extracted: List[Dict[str, Any]],
    threshold: int = 88,
) -> List[Dict[str, Any]]:
    """
    Второй проход группировки: объединяет позиции, у которых нормализованные ключи
    очень похожи (rapidfuzz token_sort_ratio >= threshold), но всё ещё разные.

    Алгоритм:
    - Собирает уникальные canonical-ключи.
    - Строит граф: рёбра между ключами с похожестью >= threshold.
    - Находит связные компоненты (Union-Find) — каждая компонента → один каноник
      (выбирается самый короткий ключ как «эталон», или первый алфавитный).
    - Перезаписывает поле «canonical» в extracted.

    Вызывается только для динамических ключей (не из каталога), чтобы не ломать
    уже правильно сопоставленные позиции.
    """
    from rapidfuzz import fuzz as _fuzz

    # Собираем все уникальные canonical-ключи.
    keys: List[str] = []
    seen: set = set()
    for row in extracted:
        c = row.get("canonical", "")
        if c and c not in seen:
            seen.add(c)
            keys.append(c)

    if len(keys) < 2:
        return extracted

    # Union-Find
    parent: Dict[str, str] = {k: k for k in keys}

    def find(x: str) -> str:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        # Выбираем «представителя» — более короткий или лексически меньший ключ.
        if len(ra) > len(rb) or (len(ra) == len(rb) and ra > rb):
            ra, rb = rb, ra
        parent[rb] = ra

    # Сравниваем все пары (O(n²), но ключей обычно < 500, быстро).
    for i in range(len(keys)):
        for j in range(i + 1, len(keys)):
            a, b = keys[i], keys[j]
            # Защита 1: первое слово (название препарата) должно совпадать.
            # Без этого «Аторис 10мг» склеится с «Дапафорс 10мг» (score ~80).
            toks_a = a.split()
            toks_b = b.split()
            first_a = next((t for t in toks_a if not t.isdigit() and t != "n"), "")
            first_b = next((t for t in toks_b if not t.isdigit() and t != "n"), "")
            if first_a != first_b:
                continue
            # Защита 2: числа (дозировки) должны совпадать.
            # Без этого «Аторис 10мг» склеится с «Аторис 20мг» (score ~90).
            import re as _re
            def _extract_nums(toks: list) -> frozenset:
                nums = set()
                for t in toks:
                    if t.isdigit():
                        nums.add(t)
                    elif "d" in t and any(c.isdigit() for c in t):  # десятичные: 12d5
                        nums.add(t)
                return frozenset(nums)
            nums_a = _extract_nums(toks_a)
            nums_b = _extract_nums(toks_b)
            if nums_a != nums_b:
                continue
            score = _fuzz.token_sort_ratio(a, b)
            if score >= threshold:
                union(a, b)

    # Строим маппинг старый_ключ → канонический_представитель.
    mapping: Dict[str, str] = {k: find(k) for k in keys}

    # Применяем маппинг.
    for row in extracted:
        c = row.get("canonical", "")
        if c in mapping:
            row["canonical"] = mapping[c]

    return extracted


def is_materia_producer_norm(text_norm: str) -> bool:
    """Проверка, что нормализованный текст относится к производителю Materia Medica."""
    if not text_norm:
        return False
    t = text_norm.replace("-", " ").replace("/", " ")
    if "материа" in t and ("медика" in t or "медиа" in t or "medica" in t):
        return True
    if "материа" in t and "холдинг" in t:
        return True
    if "materia" in t and "medica" in t:
        return True
    if "медикал" in t and "материа" in t:
        return True
    return False


def friendly_source_label(source_stem: str) -> str:
    """Человекочитаемое название дистрибьютора из имени файла."""
    s = source_stem
    mapping = {
        "верона": "Верона",
        "pharma choice": "Фарма Чойз",
        "pharm gate": "Гейт Фарм",
        "фарм гейт": "Гейт Фарм",
        "охотник": "Охотник Фарм",
        "турон": "Турон Фарм",
        "саида": "Саида Фарм",
        "фарма космос": "Фарма Космос",
        "фарма космом": "Фарма Космос",
        "гранд фарм": "Гранд Фарм",
        "аззам": "Аззам Фарм",
        "мерос": "Мерос Фарм",
        "фарм континент": "Фарм Континент",
        "фарм люкс": "Фарм Люкс",
        "гармония": "Гармония Фарм",
        "макро фарм": "Макро Фарм",
        "мемори": "Мемори Фарм",
        "релианс": "Релианс Фарм",
        "универсал ника": "Универсал Ника Фарм",
        "навкирон": "Навкирон Фарм",
        "нам экомед": "Нам экомед фарм",
        "хожи акбар": "Хожи Акбар Фарм",
        "зилол мед": "Зилол Мед",
        "гейт": "Гейт Фарм",
        "Эверест фарм": "Эверест Фарм",
        "Фарм Инвест": "Фарм Инвест",
    }
    lower = s.lower()
    for key, label in mapping.items():
        if key in lower:
            return label
    parts = []
    for p in s.replace(".", " ").replace("_", " ").split():
        if p.isdigit() or (len(p) >= 4 and p[:2].isdigit() and p[2] in ".-"):
            continue
        parts.append(p)
    return " ".join(parts) if parts else s


def build_catalog_materia_medica() -> List[Tuple[str, List[str]]]:
    """Каноническое название + синонимы для матчинга."""
    return [
        ("Анаферон детский таб №20", [
            "анаферон дет", "анаф дет", "анаферон детский таб", "анаферон детский таб n20",
            "анаферон детский таб. №20", "анаферон таб. детские №20", "анаферон таб. детск. №20",
            "анаферон таб детск", "анаферон таб детские n20",
        ]),
        ("Анаферон взрослый таб №20", ["Анаферон взрос. №20", "АНАФЕРОН №20 Д/ВЗР", "Анаферон таб. Взр №20", "анаферон таб", "Анаферон взр таб №20 !!!", "Анаферон взр таб №20", "анаферон взрослый", "анаферон таб. №20", "АНАФЕРОН ДЛЯ ВЗРОС  ТАБ N20", "анаферон для взрос", "анаферон для взрослых", "анаферон взр таб n20"]),
        ("Анаферон капли (детские) 25мл", [
            "Анаферон детский капли 25мл.", "Анаферон детск капли для п/внутрь 25мл гомеоп", "Анаферон детский капли 25мл", "Анаферон детский капли гомеоп. д/приема внутрь 25мл ", "Анаферон детский капли д/приема внутрь 25мл", "анаферон капли", "анаферон капли дет", "анаферон капли 25", "анаферон капли (детские) 25мл",
            "анаферон детский капли", "анаферон капли д/детей 25мл", "Анаферон Детский капли 25мл", "Анаферон детский капли д/детей 25мл", "Анаферон детский  капли для п/внутрь 25мл", "Анаферон детский капли д/п/внутрь 25мл"    
        ]),
        ("Эргоферон таб №20", ["эргоферон таб", "эргоферон таб. №20"]),
        ("Ренгалин таб. №20", ["ренгалин таб", "ренгалин таб. №20",]),
        ("Ренгалин раствор №1", ["Ренгалин р-р д/внутр. прим.100мл", "Ренгалин р-р. 100мл №1", "Ренгалин р-р для прием.внутрь гомеоп. 100мл", "Ренгалин р-р д/приёма внутрь 100мл ", "Ренгалин р-р 100мл №1", "ренгалин раствор","РЕНГАЛИН Р-Р Д/ПРИЕМА ВНУТРЬ ГОМЕО-Е 100МЛ", 
            "Ренгалин р-р д/пр внут гомеоп  100мл", "ренгалин 100", "ренгалин сироп", "Ренгалин р-р д/прием. внутрь 100мл", "Ренгалин р-р д/пр внут 100мл", "Ренгалин р-р д/пр. внутрь гомеоп. 100мл", "Ренгалин р-р. 100мл №1"]),
        ("Тенотен таб №40", [
            "тенотен таб. №40", "тенотен таб n40", "ТЕНОТЕН ВЗРОС  ТАБ N40",
            "тенотен взр таб", "тенотен взр", "тенотен взрос",
            "тенотен таб. д/рассас", "тенотен таб д/рассас", "тенотен таб д рассас",
        ]),
        ("Тенотен детский таб №40", [
            "Тенотен-Детский таб.№40", "тенотен дет", "тенотен таб детский", "тенотен таб. детский №40",
            "тенотен таб №40 д/детей", "тенотен таб. детс. №40", "Тенотен таб. детский №40",
        ]),
        ("Импаза №20", ["ИМПАЗА №20", "импаза таб", "импаза таб. №20"]),
        ("Афала №100", [
            "афала №100", "афала таб №100", "афала таб. №100", "афала таб n100",
            "афала гомеопат таб n100", "афала гомеопат таб №100",
        ]),
        ("Афалаза", ["афалаза"]),
        ("Климаксан таб. №20", ["климаксан таб", "климаксан таб. №20"]),
        ("Млекоин гранулы 10г", ["млекоин гранулы", "млекоин 10г", "МЛЕКОИН ГРАН.ГОМЕОП.10Г"]),
        ("Диваза", ["диваза"]),
        ("Успокой", ["успокой"]),
        ("Фарингомед", ["фарингомед"]),
    ]


def match_canonical(
    name_norm: str,
    catalog: List[Tuple[str, List[str]]],
    *,
    legacy_priority_rules: bool = True,
) -> Optional[str]:
    """По нормализованному названию возвращает каноническое или None."""
    if legacy_priority_rules:
        # Приоритетные правила для "Анаферон", чтобы капли и взрослый
        # не попадали в общий "Анаферон детский".
        if "анаферон" in name_norm:
            if ("капл" in name_norm) or ("25мл" in name_norm):
                return "Анаферон капли (детские) 25мл"
            if ("взрос" in name_norm) or ("для взрос" in name_norm) or ("д/взр" in name_norm):
                return "Анаферон взрослый таб №20"
            if "дет" in name_norm:
                return "Анаферон детский таб №20"

        if "тенотен" in name_norm:
            if ("дет" in name_norm) or ("д/дет" in name_norm):
                return "Тенотен детский таб №40"
            if ("взр" in name_norm) or ("взрос" in name_norm):
                return "Тенотен таб №40"

        # Не даём подстроке "афала" матчиться внутри "дуафалак".
        if re.search(r"\bдуафа", name_norm):
            return None

    for canonical, aliases in catalog:
        for alias in aliases:
            alias_norm = normalize_name(alias)
            if not alias_norm:
                continue
            if alias_norm in name_norm:
                return canonical
            # Защита от ложных fuzzy-совпадений (например Ревалгин -> Ренгалин).
            if fuzz.ratio(name_norm, alias_norm) >= 90 and name_norm[:5] == alias_norm[:5]:
                return canonical
        if normalize_name(canonical) in name_norm:
            return canonical
    return None


def extract_by_producer(
    df: pd.DataFrame,
    path: Path,
    producer_substrings: List[str],
    key_candidates: List[str],
    name_candidates: List[str],
    price_candidates: List[str],
    stock_candidates: List[str],
    producer_candidates: List[str],
) -> List[Dict[str, Any]]:
    """Извлекает строки, где производитель совпадает с одним из producer_substrings."""
    key_col = pick_column(list(df.columns), key_candidates)
    name_col = pick_column(list(df.columns), name_candidates)
    price_col = pick_price_column(list(df.columns), price_candidates)
    producer_col = pick_column(list(df.columns), producer_candidates)
    if not name_col or not price_col:
        raise ValueError(f"Не нашёл колонку наименования или цены в файле: {path}")
    rows = []
    for _, r in df.iterrows():
        name_val = r.get(name_col)
        if name_val is None or (isinstance(name_val, float) and pd.isna(name_val)):
            continue
        producer_val = r.get(producer_col) if producer_col else ""
        prod_norm = normalize_name(str(producer_val or ""))
        if not any(is_materia_producer_norm(prod_norm) or (s.lower() in prod_norm for s in producer_substrings)):
            continue
        price = to_number(r.get(price_col))
        rounded_price = round_price(price)
        if rounded_price is None:
            continue
        rows.append({
            "key": r.get(key_col),
            "name": name_val,
            "price": rounded_price,
            "producer": producer_val,
            "source": path.stem,
        })
    return rows


def extract_materia_any(
    df: pd.DataFrame,
    path: Path,
    key_candidates: List[str],
    name_candidates: List[str],
    price_candidates: List[str],
    stock_candidates: List[str],
    producer_candidates: List[str],
    catalog: List[Tuple[str, List[str]]],
    *,
    source_label: Optional[str] = None,
    producer_substrings: Optional[List[str]] = None,
    legacy_priority_rules: bool = True,
) -> List[Dict[str, Any]]:
    """
    Извлекает строки для матрицы: по производителю И/ИЛИ по каталогу.
    producer_substrings=None — только правило Materia Medica (как раньше).
    Иначе — подстроки в нормализованном поле производителя.
    """
    name_col = pick_column(list(df.columns), name_candidates)
    price_col = pick_price_column(list(df.columns), price_candidates)
    producer_col = pick_column(list(df.columns), producer_candidates)
    key_col = pick_column(list(df.columns), key_candidates)
    if not name_col or not price_col:
        raise ValueError(f"Не нашёл колонку наименования или цены в файле: {path}")
    rows = []
    for _, r in df.iterrows():
        name_val = r.get(name_col)
        if name_val is None or (isinstance(name_val, float) and pd.isna(name_val)):
            continue
        name_norm = normalize_name(str(name_val))
        producer_val = r.get(producer_col) if producer_col else ""
        prod_norm = normalize_name(str(producer_val or ""))
        if producer_substrings is None:
            by_producer = is_materia_producer_norm(prod_norm)
        else:
            subs = [normalize_name(s) for s in producer_substrings if str(s).strip()]
            by_producer = any(s and s in prod_norm for s in subs)
        canonical = match_canonical(
            name_norm, catalog, legacy_priority_rules=legacy_priority_rules
        )
        if not by_producer and not canonical:
            continue
        price = to_number(r.get(price_col))
        rounded_price = round_price(price)
        if rounded_price is None:
            continue
        canon_value: str
        if canonical:
            canon_value = canonical
        elif catalog:
            canon_value = str(name_val)
        else:
            canon_value = normalize_product_for_grouping(str(name_val))

        row: Dict[str, Any] = {
            "key": r.get(key_col),
            "name": name_val,
            "price": rounded_price,
            "producer": producer_val,
            "source": path.stem,
            "canonical": canon_value,
            "canonical_display": str(name_val).strip(),
        }
        if source_label:
            row["source_label"] = source_label.strip()
        rows.append(row)
    return rows


def build_matrix_from_extracted(
    extracted: List[Dict[str, Any]],
    catalog: List[Tuple[str, List[str]]],
    all_source_labels: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Строит матрицу: строки — канонические препараты, столбцы — дистрибьюторы."""
    canon_order = [c[0] for c in catalog]
    by_canon: Dict[str, Dict[str, Optional[float]]] = {c: {} for c in canon_order}
    display_by_canon: Dict[str, str] = {c: c for c in canon_order}
    sources_seen = set()
    for row in extracted:
        canon = row.get("canonical") or row.get("name")
        if isinstance(canon, str) and canon not in by_canon:
            by_canon[canon] = {}
        if isinstance(canon, str):
            disp = str(row.get("canonical_display") or canon).strip()
            if disp and canon not in display_by_canon:
                display_by_canon[canon] = disp
        if isinstance(canon, str):
            src = row.get("source_label") or friendly_source_label(row["source"])
            sources_seen.add(src)
            price = row.get("price")
            prev_price = by_canon[canon].get(src)
            if prev_price is None:
                by_canon[canon][src] = price
            elif price is not None:
                by_canon[canon][src] = min(prev_price, price)
    if all_source_labels:
        cols = ["Препарат"] + list(all_source_labels)
    else:
        cols = ["Препарат"] + sorted(sources_seen)
    dynamic_canons = [c for c in by_canon.keys() if c not in canon_order]
    if dynamic_canons:
        dynamic_canons.sort(key=lambda x: display_by_canon.get(x, x).casefold())
    final_order = canon_order + dynamic_canons
    data = []
    for canon in final_order:
        d = by_canon.get(canon, {})
        label = display_by_canon.get(canon, canon)
        row = [label] + [d.get(c) for c in cols[1:]]
        data.append(row)
    return pd.DataFrame(data, columns=cols)


def format_summary_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Приводит лист к аккуратному читаемому виду."""
    if sheet.max_row == 0 or sheet.max_column == 0:
        return

    def pixels_to_width(pixels: int) -> float:
        # Приближенный перевод пикселей в единицы ширины Excel.
        return round(max((pixels - 5) / 7, 0), 2)

    first_col_width = pixels_to_width(265)
    other_cols_width = pixels_to_width(150)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = first_col_width if col_idx == 1 else other_cols_width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            if cell.column == 1 and 2 <= cell.row <= 99:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
            cell.border = thin_border
            if cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            elif cell.column >= 2 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def format_elapsed_seconds(sec: float) -> str:
    """Человекочитаемая строка для вывода в консоль."""
    if sec >= 60:
        minutes = int(sec // 60)
        rest = sec - minutes * 60
        return f"Время работы: {minutes} мин {rest:.1f} с"
    return f"Время работы: {sec:.1f} с"


def _live_timer_caption(sec: float) -> str:
    """Короткая подпись для заголовка окна (обновляется в реальном времени)."""
    if sec >= 3600:
        h = int(sec // 3600)
        sec -= h * 3600
        m = int(sec // 60)
        s = int(sec % 60)
        return f"{h}ч {m:02d}:{s:02d}"
    if sec >= 60:
        m = int(sec // 60)
        s = int(sec % 60)
        return f"{m}:{s:02d}"
    return f"{sec:.1f} с"


def _set_console_window_title(title: str) -> None:
    """Заголовок окна консоли (сверху): не смешивается с текстом в области вывода."""
    try:
        if sys.platform == "win32":
            ctypes.windll.kernel32.SetConsoleTitleW(title)
        else:
            sys.stdout.write(f"\033]0;{title}\007")
            sys.stdout.flush()
    except Exception:
        pass


def _run_live_timer(stop: threading.Event, elapsed_fn: Callable[[], float], title_prefix: str) -> None:
    while not stop.wait(0.2):
        elapsed = elapsed_fn()
        _set_console_window_title(f"{title_prefix} — {_live_timer_caption(elapsed)}")


def start_live_console_timer(
    started: float,
    title_prefix: str = "Сводный прайс",
) -> Tuple[threading.Event, threading.Thread]:
    """Фоновый таймер в заголовке окна CMD/терминала."""
    stop = threading.Event()

    def elapsed() -> float:
        return time.perf_counter() - started

    t = threading.Thread(
        target=_run_live_timer,
        args=(stop, elapsed, title_prefix),
        daemon=True,
    )
    t.start()
    return stop, t


def finalize_live_console_timer_title(title_prefix: str, total_sec: float) -> None:
    """Заголовок после остановки таймера (строка «готово»)."""
    _set_console_window_title(f"{title_prefix} — готово, {_live_timer_caption(total_sec)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Сводный прайс из нескольких Excel")
    parser.add_argument("--config", type=Path, default=None, help="Путь к config.yml")
    parser.add_argument("--inputs", type=str, default="prices", help="Папка или файл(ы) с прайсами")
    parser.add_argument("--output", type=Path, default=None, help="Выходной Excel (по умолчанию: output/Сводный прайс + ДД.ММ.ГГГГ.xlsx)")
    parser.add_argument("--products", nargs="*", help="Список препаратов для извлечения")
    parser.add_argument("--producers", nargs="*", default=[], help="Список подстрок производителя")
    parser.add_argument("--producer-matrix", action="store_true", help="Матрица по производителю Materia Medica")
    args = parser.parse_args()
    if args.output is None:
        today = datetime.now().strftime("%d.%m.%Y")
        args.output = Path("output") / f"Сводный прайс {today}.xlsx"

    key_candidates = DEFAULT_KEY_CANDIDATES
    name_candidates = DEFAULT_NAME_CANDIDATES
    price_candidates = DEFAULT_PRICE_CANDIDATES
    stock_candidates = DEFAULT_STOCK_CANDIDATES
    producer_candidates = DEFAULT_PRODUCER_CANDIDATES

    if args.config and args.config.exists():
        with open(args.config, "r", encoding="utf-8") as f:
            cfg = yaml.safe_load(f) or {}
        cols = cfg.get("columns", {})
        key_candidates = cols.get("key_candidates", key_candidates)
        name_candidates = cols.get("name_candidates", name_candidates)
        price_candidates = cols.get("price_candidates", price_candidates)
        stock_candidates = cols.get("stock_candidates", stock_candidates)
        producer_candidates = cols.get("producer_candidates", producer_candidates)

    inputs_path = Path(args.inputs)
    if inputs_path.is_dir():
        files = list(inputs_path.glob("*.xlsx")) + list(inputs_path.glob("*.xls"))
    else:
        files = [Path(args.inputs)]

    if not files:
        print("Нет файлов для обработки.", file=sys.stderr)
        sys.exit(1)

    all_source_labels = [friendly_source_label(p.stem) for p in files]
    all_source_labels = sorted(dict.fromkeys(all_source_labels))

    catalog = build_catalog_materia_medica()

    if args.producer_matrix:
        extracted = []
        total = len(files)
        for i, path in enumerate(files, 1):
            try:
                print(f"[{i}/{total}] {path.name}", flush=True)
                df = read_price_table(
                    path,
                    key_candidates,
                    name_candidates,
                    price_candidates,
                    stock_candidates,
                    use_key=True,
                    use_name=True,
                    read_preset=None,
                )
                extracted += extract_materia_any(
                    df,
                    path,
                    key_candidates,
                    name_candidates,
                    price_candidates,
                    stock_candidates,
                    producer_candidates,
                    catalog,
                )
            except Exception as e:
                print(f"Ошибка при чтении {path}: {e}", file=sys.stderr)
        matrix = build_matrix_from_extracted(extracted, catalog, all_source_labels)
        args.output.parent.mkdir(parents=True, exist_ok=True)
        matrix.to_excel(args.output, index=False, engine="openpyxl")
        wb = openpyxl.load_workbook(args.output)
        for sheet in wb.worksheets:
            format_summary_sheet(sheet)
        wb.save(args.output)
        print(f"Сохранено: {args.output}")
        return

    all_rows = []
    for path in files:
        try:
            df = read_price_table(
                path,
                key_candidates,
                name_candidates,
                price_candidates,
                stock_candidates,
                use_key=True,
                use_name=True,
                read_preset=None,
            )
            if args.producers:
                all_rows += extract_by_producer(
                    df, path, [p.lower() for p in args.producers],
                    key_candidates, name_candidates, price_candidates, stock_candidates, producer_candidates,
                )
            else:
                for _, r in df.iterrows():
                    name_val = r.get(pick_column(list(df.columns), name_candidates))
                    price_val = r.get(pick_price_column(list(df.columns), price_candidates))
                    price_num = to_number(price_val)
                    rounded_price = round_price(price_num) if price_num is not None else None
                    if name_val is not None and rounded_price is not None:
                        all_rows.append({"name": name_val, "price": rounded_price, "source": path.stem})
        except Exception as e:
            print(f"Ошибка при чтении {path}: {e}", file=sys.stderr)
    if not all_rows:
        print("Нет данных для записи.", file=sys.stderr)
        sys.exit(1)
    out_df = pd.DataFrame(all_rows)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(args.output, index=False, engine="openpyxl")
    wb = openpyxl.load_workbook(args.output)
    for sheet in wb.worksheets:
        format_summary_sheet(sheet)
    wb.save(args.output)
    print(f"Сохранено: {args.output}")


if __name__ == "__main__":
    _timer_prefix = "Сводный прайс"
    _t0 = time.perf_counter()
    _stop_live, _live_thread = start_live_console_timer(_t0, title_prefix=_timer_prefix)
    try:
        main()
    finally:
        _stop_live.set()
        _live_thread.join(timeout=2.0)
        _total = time.perf_counter() - _t0
        finalize_live_console_timer_title(_timer_prefix, _total)
        print(format_elapsed_seconds(_total), flush=True)
