# -*- coding: utf-8 -*-
"""
Анализ сводного прайса: для заданного набора препаратов — лучший оптовик по каждой позиции
и сводка по «корзине» (сумма минимально доступных цен у одного поставщика).
Цены в прайсах и отчёте — в узбекских сумах (UZS).
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, DefaultDict, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from cheapest_svodny_prices import (
    PRODUCT_COL_CANDIDATES,
    SUPPLIER_COL_CANDIDATES,
    collect_files,
    find_column_index,
    normalize_text,
    round_price,
    to_number,
)

# Доп. варианты шапки в выгрузках (первая строка не всегда «Препарат»).
_EXTRA_PRODUCT_HEADERS = [
    "Препараты",
    "Торговое наименование",
    "Наименование товара",
]
PRODUCT_HEADERS_FOR_MATCH: List[str] = list(dict.fromkeys(PRODUCT_COL_CANDIDATES + _EXTRA_PRODUCT_HEADERS))


# Подпись валюты в тексте отчёта (цены в сводном прайсе — узбекский сум).
CURRENCY_SUFFIX = "сум"

# Канонические названия и порядок вывода (как в запросе пользователя).
TARGET_PREP_NAMES: List[str] = [
    "Анаферон детский",
    "Анаферон детский капли",
    "Эргоферон",
    "Ренгалин",
    "Тенотен",
    "Тенотен детский",
]


def min_positions_for_basket_report() -> int:
    """Минимум позиций с ценой у оптовика, чтобы попасть в таблицу «корзина» (сейчас: все минус одна)."""
    return max(0, len(TARGET_PREP_NAMES) - 1)


def classify_preparation(raw_name: str) -> Optional[str]:
    """
    Сопоставляет строку из сводного прайса с одной из целевых позиций.
    Учитываются типичные варианты написания в наименованиях.
    """
    n = normalize_text(raw_name)
    if not n:
        return None

    if "анаферон" in n:
        if "капл" in n:
            return "Анаферон детский капли"
        if "дет" in n:
            return "Анаферон детский"
        return None

    if "эргоферон" in n:
        return "Эргоферон"

    if "ренгалин" in n:
        return "Ренгалин"

    if "тенотен" in n:
        if "дет" in n:
            return "Тенотен детский"
        return "Тенотен"

    # Канонические подписи в сводной матрице (таб №, капли 25мл и т.д.): ищем длинные цели первыми.
    for t in sorted(TARGET_PREP_NAMES, key=len, reverse=True):
        tn = normalize_text(t)
        if len(tn) >= 4 and tn in n:
            return t

    return None


def _headers_list(df: pd.DataFrame) -> List[object]:
    return [df.columns[i] for i in range(len(df.columns))]


def _merge_price(
    store: DefaultDict[str, Dict[str, int]],
    prep: str,
    supplier: str,
    price: Optional[int],
) -> None:
    if price is None or not supplier:
        return
    s = str(supplier).strip()
    if not s:
        return
    prev = store[prep].get(s)
    if prev is None or price < prev:
        store[prep][s] = price


def ingest_horizontal_sheet(
    df: pd.DataFrame,
    store: DefaultDict[str, Dict[str, int]],
) -> None:
    headers = _headers_list(df)
    prod_col_idx = find_column_index(headers, PRODUCT_HEADERS_FOR_MATCH)
    if prod_col_idx is None:
        return
    prod_col = headers[prod_col_idx - 1]
    supplier_cols = [c for c in df.columns if c != prod_col]

    for _, row in df.iterrows():
        raw = row.get(prod_col)
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        prep = classify_preparation(str(raw))
        if prep is None:
            continue
        for sup_col in supplier_cols:
            price = round_price(to_number(row.get(sup_col)))
            _merge_price(store, prep, str(sup_col), price)


def ingest_vertical_sheet(
    df: pd.DataFrame,
    store: DefaultDict[str, Dict[str, int]],
) -> None:
    headers = _headers_list(df)
    sup_col_idx = find_column_index(headers, SUPPLIER_COL_CANDIDATES)
    if sup_col_idx is None:
        return
    sup_col = headers[sup_col_idx - 1]
    product_cols = [c for c in df.columns if c != sup_col]

    for prod_col in product_cols:
        raw_header = str(prod_col).strip() if prod_col is not None else ""
        if not raw_header:
            continue
        prep = classify_preparation(raw_header)
        if prep is None:
            continue
        for _, row in df.iterrows():
            supplier = row.get(sup_col)
            if supplier is None or (isinstance(supplier, float) and pd.isna(supplier)):
                continue
            price = round_price(to_number(row.get(prod_col)))
            _merge_price(store, prep, str(supplier).strip(), price)


def ingest_dataframe(
    df: pd.DataFrame,
    store: DefaultDict[str, Dict[str, int]],
) -> None:
    headers_norm = {normalize_text(h) for h in _headers_list(df)}
    is_horizontal = any(normalize_text(c) in headers_norm for c in PRODUCT_HEADERS_FOR_MATCH)
    is_vertical = any(normalize_text(c) in headers_norm for c in SUPPLIER_COL_CANDIDATES)
    if is_horizontal:
        ingest_horizontal_sheet(df, store)
    elif is_vertical:
        ingest_vertical_sheet(df, store)


def _target_price_count(store: DefaultDict[str, Dict[str, int]]) -> int:
    return sum(1 for p in TARGET_PREP_NAMES if store.get(p))


def _best_header_for_sheet(xl: pd.ExcelFile, sheet: str) -> Optional[int]:
    """Номер строки заголовка (0 — первая строка листа), дающий максимум целевых позиций с ценой."""
    best_h: Optional[int] = None
    best_cnt = 0
    for h in range(0, 6):
        try:
            df = pd.read_excel(xl, sheet_name=sheet, header=h)
        except Exception:
            continue
        if df.empty or len(df.columns) < 2:
            continue
        trial: DefaultDict[str, Dict[str, int]] = defaultdict(dict)
        ingest_dataframe(df, trial)
        cnt = _target_price_count(trial)
        if cnt > best_cnt:
            best_cnt = cnt
            best_h = h
    return best_h


def collect_prices_from_workbook(path: Path) -> DefaultDict[str, Dict[str, int]]:
    store: DefaultDict[str, Dict[str, int]] = defaultdict(dict)
    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
    except Exception as exc:
        print(f"Не удалось открыть {path}: {exc}", file=sys.stderr)
        return store

    for sheet in xl.sheet_names:
        header_row = _best_header_for_sheet(xl, sheet)
        if header_row is None:
            continue
        try:
            df = pd.read_excel(xl, sheet_name=sheet, header=header_row)
        except Exception as exc:
            print(f"Лист «{sheet}» в {path.name}: пропуск ({exc})", file=sys.stderr)
            continue
        if df.empty:
            continue
        ingest_dataframe(df, store)
    return store


def best_per_preparation(
    store: DefaultDict[str, Dict[str, int]],
) -> List[Tuple[str, Optional[str], Optional[int], int]]:
    """
    Возвращает для каждой целевой позиции: (название, лучший оптовик, лучшая цена, число поставщиков с ценой).
    """
    rows: List[Tuple[str, Optional[str], Optional[int], int]] = []
    for prep in TARGET_PREP_NAMES:
        prices = dict(store.get(prep, {}))
        if not prices:
            rows.append((prep, None, None, 0))
            continue
        best_sup, best_p = min(prices.items(), key=lambda x: x[1])
        rows.append((prep, best_sup, best_p, len(prices)))
    return rows


def missing_for_supplier(store: DefaultDict[str, Dict[str, int]], supplier: str) -> List[str]:
    """Препараты из целевого списка, по которым у оптовика нет цены."""
    return [
        p
        for p in TARGET_PREP_NAMES
        if store.get(p, {}).get(supplier) is None
    ]


def basket_totals(
    store: DefaultDict[str, Dict[str, int]],
) -> List[Tuple[str, int, int, int, str]]:
    """
    Для каждого оптовика:
    (имя, сумма по доступным, число позиций с ценой, всего целевых, строка «отсутствуют» или пусто).
    Сортировка по сумме (меньше — выгоднее для закупки всего у одного).
    """
    all_suppliers = set()
    for prep in TARGET_PREP_NAMES:
        all_suppliers.update(store.get(prep, {}).keys())

    total_targets = len(TARGET_PREP_NAMES)
    out: List[Tuple[str, int, int, int, str]] = []
    for sup in sorted(all_suppliers):
        ssum = 0
        count = 0
        for prep in TARGET_PREP_NAMES:
            p = store.get(prep, {}).get(sup)
            if p is not None:
                ssum += p
                count += 1
        if count > 0:
            miss = missing_for_supplier(store, sup)
            miss_str = "; ".join(miss) if miss else ""
            out.append((sup, ssum, count, total_targets, miss_str))
    out.sort(key=lambda x: (x[1], -x[2], x[0].lower()))
    return out


def default_inputs_dir() -> Path:
    try:
        from app_settings import output_dir

        p = output_dir()
        if p.is_dir() and any(p.glob("*.xlsx")):
            return p
    except Exception:
        pass
    return Path("output")


def resolve_input_path(file_arg: Optional[Path], inputs_dir: Optional[Path]) -> Path:
    if file_arg is not None and file_arg.is_file():
        return file_arg
    root = inputs_dir if inputs_dir is not None else default_inputs_dir()
    if not root.is_dir():
        raise FileNotFoundError(f"Папка не найдена: {root}")
    files = collect_files(root)
    if not files:
        raise FileNotFoundError(
            f"В {root} нет подходящих XLSX (ожидается сводный прайс)."
        )
    return max(files, key=lambda p: p.stat().st_mtime)


def print_report(
    path: Path,
    per_prep: List[Tuple[str, Optional[str], Optional[int], int]],
    baskets: List[Tuple[str, int, int, int, str]],
) -> None:
    print(f"Файл: {path}", flush=True)
    print("\n—— Лучшая цена по каждому препарату ——", flush=True)
    for prep, sup, price, n in per_prep:
        if sup is None or price is None:
            print(f"  • {prep}: нет цен в файле", flush=True)
        else:
            extra = f" (вариантов у оптовиков: {n})" if n > 1 else ""
            print(f"  • {prep}: {sup} — {price} {CURRENCY_SUFFIX}{extra}", flush=True)

    n_tar = len(TARGET_PREP_NAMES)
    min_show = min_positions_for_basket_report()
    print(
        "\n—— Закупка у одного оптовика "
        f"(в списке только поставщики с ценой минимум по {min_show} из {n_tar}; сумма по доступным ценам) ——",
        flush=True,
    )
    if not baskets:
        print(
            f"  Нет поставщиков с ценой минимум по {min_show} из {n_tar} (остальные отсеяны).",
            flush=True,
        )
        return
    complete = [b for b in baskets if b[2] == b[3]]
    if complete:
        best = complete[0]
        print(
            f"  Лучший по полной корзине ({best[2]}/{best[3]} позиций): {best[0]} — {best[1]} {CURRENCY_SUFFIX}",
            flush=True,
        )
    else:
        print(
            f"  Ни у кого нет цен сразу по всем {n_tar} позициям; в списке — поставщики с {min_show}/{n_tar}.",
            flush=True,
        )
    for sup, ssum, cnt, tot, miss in baskets:
        tag = " (полный набор)" if cnt == tot else f" ({cnt}/{tot} поз.)"
        miss_part = f" — нет: {miss}" if miss else ""
        print(f"  • {sup}: сумма {ssum} {CURRENCY_SUFFIX}{tag}{miss_part}", flush=True)


def write_report_excel(
    out_path: Path,
    per_prep: List[Tuple[str, Optional[str], Optional[int], int]],
    baskets: List[Tuple[str, int, int, int, str]],
    source_file: Path,
) -> None:
    rows = []
    for prep, sup, price, n in per_prep:
        rows.append(
            {
                "Препарат": prep,
                "Лучший оптовик": sup or "",
                "Цена (сум, UZS)": price if price is not None else "",
                "Число оптовиков с ценой": n,
            }
        )
    df1 = pd.DataFrame(rows)
    rows2 = []
    for sup, ssum, cnt, tot, miss in baskets:
        rows2.append(
            {
                "Оптовик": sup,
                "Сумма по доступным (сум, UZS)": ssum,
                "Позиций с ценой": cnt,
                "Всего целевых позиций": tot,
                "Отсутствуют препараты": miss,
            }
        )
    df2 = pd.DataFrame(rows2)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="По препаратам", index=False)
        df2.to_excel(writer, sheet_name="Корзина у одного", index=False)
        n_tar = len(TARGET_PREP_NAMES)
        min_show = min_positions_for_basket_report()
        meta = pd.DataFrame(
            [
                {
                    "Исходный файл": str(source_file.resolve()),
                    "Валюта": "узбекский сум (UZS)",
                    "Отбор «Корзина у одного»": (
                        f"только оптовики с ценой минимум по {min_show} из {n_tar} препаратов"
                    ),
                }
            ],
        )
        meta.to_excel(writer, sheet_name="Источник", index=False)
    _autofit_workbook_columns(out_path)


def _cell_text_width_units(value: object) -> float:
    """Оценка «ширины» текста в условных знаках Excel (латиница уже кириллицы)."""
    s = str(value if value is not None else "").replace("\n", " ")
    if not s:
        return 0.0
    units = 0.0
    for ch in s:
        o = ord(ch)
        units += 1.0 if o < 128 else 1.85
    return units


def _autofit_workbook_columns(path: Path) -> None:
    """Подгонка ширины столбцов под содержимое на всех листах."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_u = 0.0
            for row_idx in range(1, ws.max_row + 1):
                max_u = max(max_u, _cell_text_width_units(ws.cell(row=row_idx, column=col_idx).value))
            width = max(8.43, min(max_u + 2.0, 255.0))
            ws.column_dimensions[letter].width = width
    wb.save(path)


def default_report_path(directory: Path) -> Path:
    """Имя файла: Анализ оптовиков ДД.ММ.ГГГГ (N препаратов).xlsx"""
    today = datetime.now().strftime("%d.%m.%Y")
    n = len(TARGET_PREP_NAMES)
    return directory / f"Анализ оптовиков {today} ({n} препаратов).xlsx"


def run_analysis_to_excel(
    *,
    source_file: Optional[Path] = None,
    inputs_dir: Optional[Path] = None,
    output_file: Optional[Path] = None,
    report_directory: Optional[Path] = None,
    progress: Optional[Callable[[str], None]] = None,
    emit_console_report: bool = False,
) -> Path:
    """
    Строит отчёт по сводному прайсу (горизонтальному или вертикальному).
    - output_file: явный путь к отчёту;
    - иначе report_directory + автоматическое имя;
    - иначе папка рядом с исходным XLSX (режим командной строки).
    Если задан только inputs_dir (например папка output приложения), берётся последний подходящий сводный файл.
    """
    def _log(msg: str) -> None:
        if progress:
            progress(msg)

    try:
        path = resolve_input_path(source_file, inputs_dir)
    except FileNotFoundError as e:
        raise ValueError(str(e)) from e

    _log(f"Источник: {path.name}")
    store = collect_prices_from_workbook(path)
    if not any(store.get(p) for p in TARGET_PREP_NAMES):
        raise ValueError(
            "В файле не найдено ни одной из целевых позиций с распознанной ценой. "
            "Проверьте: 1) это сводная матрица (препараты × оптовики) из svodny_price; "
            "2) в ячейках числа, а не текст; 3) файл сохранён в Excel с пересчётом формул. "
            "Список целевых препаратов задан в TARGET_PREP_NAMES в best_suppliers_analysis.py."
        )

    per_prep = best_per_preparation(store)
    min_pos = min_positions_for_basket_report()
    baskets = [b for b in basket_totals(store) if b[2] >= min_pos]

    if output_file is not None:
        out = output_file
    elif report_directory is not None:
        out = default_report_path(report_directory)
    else:
        out = default_report_path(path.parent)

    write_report_excel(out, per_prep, baskets, path)
    if emit_console_report:
        print_report(path, per_prep, baskets)
    _log(f"Отчёт: {out.name}")
    return out


def main() -> None:
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    parser = argparse.ArgumentParser(
        description="Анализ сводного прайса (цены в узбекских сумах UZS): лучшие оптовики для выбранных препаратов",
    )
    parser.add_argument("--file", type=Path, default=None, help="Путь к XLSX сводного прайса")
    parser.add_argument(
        "--inputs-dir",
        type=Path,
        default=None,
        help="Папка с выгрузкой (берётся последний «Сводный прайс*.xlsx»)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help='Сохранить отчёт в XLSX (по умолчанию: «Анализ оптовиков ДД.ММ.ГГГГ (6 препаратов).xlsx» рядом с исходным файлом)',
    )
    args = parser.parse_args()

    try:
        out = run_analysis_to_excel(
            source_file=args.file,
            inputs_dir=args.inputs_dir,
            output_file=args.output,
            report_directory=args.output.parent if args.output else None,
            progress=lambda m: print(m, flush=True),
            emit_console_report=True,
        )
        print(f"\nОтчёт Excel: {out}", flush=True)
    except ValueError as e:
        msg = str(e)
        print(msg, file=sys.stderr)
        low = msg.lower()
        if "папка не найдена" in low or "нет подходящих" in low:
            sys.exit(1)
        if "не найдено ни одной из целевых позиций" in low:
            sys.exit(3)
        sys.exit(2)


if __name__ == "__main__":
    main()
