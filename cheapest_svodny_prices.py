"""Подсветка минимальной цены по препарату в сводных прайсах output."""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

PRODUCT_COL_CANDIDATES = [
    "Препарат",
    "Наименование",
    "Наименование препарата",
    "Наименование препаратов",
    "Позиция",
    "Товар",
]

SUPPLIER_COL_CANDIDATES = [
    "Оптовик",
    "Поставщик",
    "Дистрибьютор",
]


_STAR_MARKERS_RE = re.compile(r"[\*＊⁎∗⋆]+")
_TM_RE = re.compile(r"[®™©]+")


def normalize_text(value: object) -> str:
    """Для сравнения ячеек и препаратов; убирает *** и аналоги из прайсов (напр. Верона)."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = _STAR_MARKERS_RE.sub("", s)
    return _TM_RE.sub("", s)


def find_column_index(headers: List[object], candidates: List[str]) -> Optional[int]:
    by_norm = {normalize_text(header): idx for idx, header in enumerate(headers, start=1)}
    for cand in candidates:
        idx = by_norm.get(normalize_text(cand))
        if idx is not None:
            return idx
    return None


def to_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\xa0", "").replace(" ", "")
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def round_price(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    return int(round(value))


def collect_records_from_sheet(path: Path, sheet_name: str) -> List[Tuple[str, int, str, int, int]]:
    records: List[Tuple[str, int, str, int, int]] = []
    try:
        wb = load_workbook(path)
    except Exception as exc:
        print(f"Пропуск {path.name}: не удалось открыть файл ({exc})", file=sys.stderr)
        return records

    if sheet_name not in wb.sheetnames:
        return records

    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
    headers_norm = {normalize_text(h) for h in headers}

    is_horizontal = any(normalize_text(c) in headers_norm for c in PRODUCT_COL_CANDIDATES)
    is_vertical = any(normalize_text(c) in headers_norm for c in SUPPLIER_COL_CANDIDATES)

    if is_horizontal:
        product_col = find_column_index(headers, PRODUCT_COL_CANDIDATES)
        if product_col is None:
            return records
        supplier_cols = [c for c in range(1, ws.max_column + 1) if c != product_col]
        for row_idx in range(2, ws.max_row + 1):
            product = str(ws.cell(row=row_idx, column=product_col).value or "").strip()
            if not product:
                continue
            for col_idx in supplier_cols:
                price = round_price(to_number(ws.cell(row=row_idx, column=col_idx).value))
                if price is None:
                    continue
                records.append((normalize_text(product), price, product, row_idx, col_idx))
        return records

    if is_vertical:
        supplier_col = find_column_index(headers, SUPPLIER_COL_CANDIDATES)
        if supplier_col is None:
            return records
        product_cols = [c for c in range(1, ws.max_column + 1) if c != supplier_col]
        for col_idx in product_cols:
            product = str(ws.cell(row=1, column=col_idx).value or "").strip()
            if not product:
                continue
            for row_idx in range(2, ws.max_row + 1):
                price = round_price(to_number(ws.cell(row=row_idx, column=col_idx).value))
                if price is None:
                    continue
                records.append((normalize_text(product), price, product, row_idx, col_idx))
    return records


def collect_records(path: Path) -> List[Tuple[str, int, str, str, int, int]]:
    out: List[Tuple[str, int, str, str, int, int]] = []
    try:
        wb = load_workbook(path, read_only=True)
    except Exception as exc:
        print(f"Пропуск {path.name}: не удалось прочитать файл ({exc})", file=sys.stderr)
        return out

    for sheet_name in wb.sheetnames:
        rows = collect_records_from_sheet(path, sheet_name)
        for key, price, product, row_idx, col_idx in rows:
            out.append((key, price, product, sheet_name, row_idx, col_idx))
    return out


def paint_min_cells(
    path: Path,
    records: List[Tuple[str, int, str, str, int, int]],
    min_price_by_product: Dict[str, int],
) -> int:
    try:
        wb = load_workbook(path)
    except Exception as exc:
        print(f"Пропуск {path.name}: не удалось открыть файл для записи ({exc})", file=sys.stderr)
        return 0

    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    painted = 0

    for key, price, _product, sheet_name, row_idx, col_idx in records:
        if key not in min_price_by_product:
            continue
        if price != min_price_by_product[key]:
            continue
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.cell(row=row_idx, column=col_idx).fill = green_fill
        painted += 1

    if painted > 0:
        wb.save(path)
    return painted


def collect_files(inputs_dir: Path) -> List[Path]:
    files = sorted(inputs_dir.glob("*.xlsx"))
    svodny = [
        p for p in files
        if "сводный прайс" in p.stem.lower()
    ]
    return svodny if svodny else files


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Подсвечивает минимальную цену по препарату в сводных прайсах"
    )
    parser.add_argument("--inputs", type=Path, default=Path("output"), help="Папка со сводными прайсами")
    args = parser.parse_args()

    if not args.inputs.exists() or not args.inputs.is_dir():
        print(f"Папка не найдена: {args.inputs}", file=sys.stderr)
        sys.exit(1)

    files = collect_files(args.inputs)
    if not files:
        print(f"В папке {args.inputs} не найдено XLSX-файлов.", file=sys.stderr)
        sys.exit(1)

    all_records: Dict[Path, List[Tuple[str, int, str, str, int, int]]] = {}
    min_price_by_product: Dict[str, int] = {}

    for path in files:
        file_records = collect_records(path)
        all_records[path] = file_records
        for key, price, _product, _sheet_name, _row_idx, _col_idx in file_records:
            prev = min_price_by_product.get(key)
            if prev is None or price < prev:
                min_price_by_product[key] = price

    if not min_price_by_product:
        print("Не удалось извлечь цены ни из одного файла.", file=sys.stderr)
        sys.exit(1)

    total_painted = 0
    changed_files = 0
    for path in files:
        painted = paint_min_cells(path, all_records.get(path, []), min_price_by_product)
        if painted > 0:
            changed_files += 1
            total_painted += painted

    print(f"Готово. Подсвечено ячеек: {total_painted}. Изменено файлов: {changed_files}.")


def highlight_minimum_prices(inputs_dir: Path) -> str:
    """
    Подсветка минимальных цен во всех подходящих XLSX в папке.
    Возвращает текстовое резюме; бросает ValueError при отсутствии данных/папки.
    """
    if not inputs_dir.exists() or not inputs_dir.is_dir():
        raise ValueError(f"Папка не найдена: {inputs_dir}")
    files = collect_files(inputs_dir)
    if not files:
        raise ValueError(f"В папке {inputs_dir} не найдено XLSX-файлов.")
    all_records: Dict[Path, List[Tuple[str, int, str, str, int, int]]] = {}
    min_price_by_product: Dict[str, int] = {}
    for path in files:
        file_records = collect_records(path)
        all_records[path] = file_records
        for key, price, _product, _sheet_name, _row_idx, _col_idx in file_records:
            prev = min_price_by_product.get(key)
            if prev is None or price < prev:
                min_price_by_product[key] = price
    if not min_price_by_product:
        raise ValueError("Не удалось извлечь цены ни из одного файла.")
    total_painted = 0
    changed_files = 0
    for path in files:
        painted = paint_min_cells(path, all_records.get(path, []), min_price_by_product)
        if painted > 0:
            changed_files += 1
            total_painted += painted
    return f"Подсвечено ячеек: {total_painted}. Изменено файлов: {changed_files}."


if __name__ == "__main__":
    main()
